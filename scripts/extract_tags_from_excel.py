import argparse
import re
from pathlib import Path
import openpyxl


PATTERN = re.compile(r"(?i)\b[A-Z0-9_]+\.[A-Z0-9_-]+(?:-[0-9]{2})?\.[A-Z0-9_.-]+\.(?:PV|OMR|AV|EU|VAL)\b")


def extract_tags(xlsx: Path, sheet: str | None = None, limit: int | None = None) -> list[str]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    tags: set[str] = set()
    for name in sheets:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    for m in PATTERN.finditer(v.strip()):
                        tags.add(m.group(0))
                        if limit and len(tags) >= limit:
                            wb.close()
                            return sorted(tags)
    wb.close()
    return sorted(tags)


def main():
    ap = argparse.ArgumentParser(description="Extract PI tag-like strings from an Excel workbook")
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--sheet", type=str, default=None)
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    tags = extract_tags(args.xlsx, sheet=args.sheet, limit=args.limit)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(tags), encoding="utf-8")
    print(f"Found {len(tags)} tags. Wrote: {args.out}")
    if tags:
        print("Sample:", tags[:10])


if __name__ == "__main__":
    main()

