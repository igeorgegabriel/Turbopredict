#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import argparse
import re
import openpyxl

PI_FUNCS = ("PICurrVal", "PISampDat", "PICompDat", "PISummary", "PIData")
PAT = re.compile(r"^\s*=?\s*(?:" + "|".join(PI_FUNCS) + r")\s*\(\s*\"([^\"]+)\"", re.IGNORECASE)

def _col_to_idx(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column: {col}")
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def extract_column(xlsx: Path, col: str = 'E', contains: str | None = None) -> list[str]:
    # Use read_only=False so openpyxl exposes formulas reliably
    wb = openpyxl.load_workbook(xlsx, data_only=False, read_only=False)
    try:
        out: list[str] = []
        seen: set[str] = set()
        cidx = _col_to_idx(col)
        for name in wb.sheetnames:
            ws = wb[name]
            max_row = ws.max_row or 0
            for r in range(2, max_row + 1):
                c = ws.cell(row=r, column=cidx)
                v = c.value
                if not isinstance(v, str):
                    continue
                s = v.strip()
                if not s or s[0] != '=':
                    continue
                m = PAT.match(s)
                if not m:
                    continue
                path = m.group(1).strip()
                if contains and contains not in path:
                    continue
                if path not in seen:
                    seen.add(path)
                    out.append(path)
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract AF attribute paths from column E formulas across all sheets")
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--contains", default=None, help="Only include formulas whose path contains this substring")
    ap.add_argument("--col", default='E', help="Column letter to scan for formulas (default E)")
    args = ap.parse_args()

    tags = extract_column(args.xlsx, col=args.col, contains=args.contains)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(tags) + ("\n" if tags else ""), encoding="utf-8")
    print(f"Extracted {len(tags)} AF path(s) -> {args.out}")
    if tags:
        print("Sample:", tags[:5])
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
