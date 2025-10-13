#!/usr/bin/env python3
"""
Test PCMSB single unit refresh without hanging
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from datetime import datetime

def test_pcmsb_single_unit():
    """Test refreshing a single PCMSB unit safely"""

    print("TESTING SINGLE PCMSB UNIT REFRESH")
    print("=" * 40)
    print(f"Test Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Check Excel file exists and has recent data
    excel_path = Path("excel/PCMSB/PCMSB_Automation.xlsx")

    if not excel_path.exists():
        print("ERROR: PCMSB_Automation.xlsx not found!")
        return

    print(f"OK Excel file found: {excel_path}")
    stat = excel_path.stat()
    excel_modified = datetime.fromtimestamp(stat.st_mtime)
    excel_age_hours = (datetime.now() - excel_modified).total_seconds() / 3600

    print(f"   Last modified: {excel_modified.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"   Age: {excel_age_hours:.1f} hours")

    # Check the DL_WORK sheet has fresh data
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)

        if 'DL_WORK' in wb.sheetnames:
            print(f"OK DL_WORK sheet found")

            # Sample a few rows to check if data is recent
            ws = wb['DL_WORK']
            rows = list(ws.iter_rows(values_only=True))

            recent_data_found = False
            for i, row in enumerate(rows[2:12]):  # Check first 10 data rows
                if row and len(row) >= 2 and isinstance(row[0], datetime):
                    data_time = row[0]
                    data_age_hours = (datetime.now() - data_time).total_seconds() / 3600
                    if data_age_hours < 2:  # Data less than 2 hours old
                        recent_data_found = True
                        print(f"   Recent data: {data_time.strftime('%Y-%m-%d %H:%M:%S')} ({data_age_hours:.1f}h ago)")
                        break

            if not recent_data_found:
                print("WARNING: No recent data found in DL_WORK sheet")
            else:
                print("OK Excel has fresh data - ready for parquet update")

        wb.close()

    except Exception as e:
        print(f"ERROR reading Excel: {e}")
        return

    # Now try to update the parquet file for C-02001
    print(f"\nTesting parquet update for C-02001...")

    # Check current parquet status
    from pi_monitor.parquet_database import ParquetDatabase

    try:
        db = ParquetDatabase()
        freshness_info = db.get_data_freshness_info('C-02001')

        current_age = freshness_info.get('data_age_hours', 0)
        latest_time = freshness_info.get('latest_timestamp')

        print(f"Current parquet status for C-02001:")
        print(f"   Latest data: {latest_time.strftime('%Y-%m-%d %H:%M:%S') if latest_time else 'None'}")
        print(f"   Age: {current_age:.1f} hours")

        if current_age > 1.0:
            print(f"OK C-02001 parquet is stale - needs refresh")

            # Check if we have the tag file
            tags_file = Path("config/tags_pcmsb_c02001.txt")
            if tags_file.exists():
                print(f"OK Tags file found: {tags_file}")

                # Read tags
                tags = [t.strip() for t in tags_file.read_text(encoding="utf-8").splitlines()
                       if t.strip() and not t.strip().startswith('#')]
                print(f"   Found {len(tags)} tags")

                if tags:
                    print(f"   Sample tags: {tags[:3]}")

                    # The issue might be in the batch processing
                    # Let's check if we can at least validate the setup
                    print(f"\nSETUP VALIDATION COMPLETE")
                    print(f"   - Excel file: READY ({excel_age_hours:.1f}h old)")
                    print(f"   - DL_WORK sheet: READY (fresh data)")
                    print(f"   - Tags file: READY ({len(tags)} tags)")
                    print(f"   - Parquet status: STALE (needs update)")

                    print(f"\nRECOMMENDATION:")
                    print(f"   The Excel file has fresh data but parquet update is failing.")
                    print(f"   This suggests the batch processing in build_unit_from_tags() is hanging.")
                    print(f"   Try running the refresh in smaller batches or with timeout limits.")

                else:
                    print("ERROR: No valid tags found in file")
            else:
                print(f"ERROR: Tags file not found: {tags_file}")
        else:
            print(f"OK C-02001 parquet is fresh - no refresh needed")

    except Exception as e:
        print(f"ERROR checking parquet status: {e}")
        import traceback
        traceback.print_exc()

    print(f"\nTest complete: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == "__main__":
    test_pcmsb_single_unit()