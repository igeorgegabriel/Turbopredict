#!/usr/bin/env python3
"""
Practical example of fetching ABF PI data using the configured Excel file.
Demonstrates both reading existing data and fetching new data.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys
from datetime import datetime, timedelta

def demonstrate_abf_data_fetching():
    """Show how ABF PI data fetching works with the current setup."""
    
    excel_path = "excel/ABFSB/ABFSB_Automation_Master.xlsx"
    
    print("=" * 60)
    print("ABF PI DATA FETCHING DEMONSTRATION")
    print("=" * 60)
    print(f"Excel File: {excel_path}")
    print(f"Absolute Path: C:\\Users\\george.gabrielujai\\Documents\\CodeX\\{excel_path}")
    
    # 1. Show current data structure
    print("\n1. CURRENT DATA IN SHEET1:")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']
    
    # Get column headers
    headers = [ws.cell(1, col).value for col in range(1, 7)]
    print(f"   Columns: {headers}")
    
    # Show sample data
    print(f"   Sample data (first 3 rows):")
    for row in range(2, 5):
        values = [ws.cell(row, col).value for col in range(1, 4)]  # First 3 columns
        print(f"   Row {row}: {values}")
    
    wb.close()
    
    # 2. Show how to fetch new data
    print("\n2. FETCHING NEW ABF DATA:")
    print("   The system uses PI DataLink formulas in Excel:")
    print("   Formula: =PISampDat(\"ABF.07-MT001.FI-07054.PV\", \"-24h\", \"*\", \"-0.1h\", 1, \"\\\\PTSG-1MMPDPdb01\")")
    
    # 3. Available ABF tags
    print("\n3. AVAILABLE ABF TAGS:")
    tags_file = "config/tags_abf_07mt01_k001.txt"
    with open(tags_file, 'r') as f:
        tags = [line.strip() for line in f if line.strip() and not line.startswith('#')]
        print(f"   Total tags configured: {len(tags)}")
        print(f"   Sample tags:")
        for i, tag in enumerate(tags[:5]):
            print(f"     {i+1}. {tag}")
        if len(tags) > 5:
            print(f"     ... and {len(tags)-5} more tags")
    
    # 4. Ready-to-use commands
    print("\n4. READY-TO-USE COMMANDS:")
    print("   Fetch all ABF data:")
    print("   > python scripts/build_abf_07mt01_k001.py")
    print("")
    print("   Fetch specific ABF tag manually:")
    print("   > python -c \"")
    print("   from pi_monitor.batch import build_unit_from_tags")
    print("   from pathlib import Path")
    print("   build_unit_from_tags(")
    print("       Path('excel/ABFSB/ABFSB_Automation_Master.xlsx'),")
    print("       ['ABF.07-MT001.FI-07054.PV'],")
    print("       Path('data/processed/abf_sample.parquet'),")
    print("       plant='ABF', unit='07-MT01-K001',")
    print("       server=r'\\\\PTSG-1MMPDPdb01')")
    print("   \"")
    
    return 0

if __name__ == "__main__":
    sys.exit(demonstrate_abf_data_fetching())