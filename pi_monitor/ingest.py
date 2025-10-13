from __future__ import annotations

from pathlib import Path
import os
import time
import re
import pandas as pd


def _detect_header_row(raw: pd.DataFrame, search_rows: int = 10) -> int | None:
    """Return index of header row containing 'TIME' within first search_rows."""
    for i in range(min(search_rows, len(raw))):
        vals = [str(v).strip().upper() if isinstance(v, str) else v for v in raw.iloc[i].tolist()]
        if "TIME" in vals:
            return i
    return None



def _try_parse_time_value_block(raw: pd.DataFrame, max_scan_cols: int = 64) -> pd.DataFrame | None:
    """Heuristic parser for sheets without a 'TIME' header.

    Scans the left-most columns (up to max_scan_cols) for adjacent columns where the
    left looks like timestamps (datetime strings or Excel serials) and the right is
    numeric. Returns a tidy DataFrame with columns ['time','value'] if successful.
    """
    if raw is None or raw.empty:
        return None

    import warnings as _w
    with _w.catch_warnings():
        _w.simplefilter("ignore")
        # Limit to a reasonable area (first rows/cols)
        ncols = min(max_scan_cols, raw.shape[1])
        best = None
        best_count = 0
        for i in range(max(0, ncols - 1)):
            c_time = i
            c_val = i + 1
            tser = raw.iloc[:, c_time]
            vser = raw.iloc[:, c_val]

            # Try string datetime first
            t1 = pd.to_datetime(tser, errors='coerce')
            # Fallback for Excel serial numbers
            try:
                num = pd.to_numeric(tser, errors='coerce')
                t2 = pd.to_datetime(num, unit='d', origin='1899-12-30')
            except Exception:
                t2 = pd.Series([pd.NaT] * len(tser))
            # Choose the conversion with more non-null, but prefer Excel serial over string for numeric data
            t1_count = t1.notna().sum()
            t2_count = t2.notna().sum()
            if t2_count > 0 and t2_count >= t1_count * 0.9:  # Prefer Excel serial if it has >=90% of string conversion count
                t = t2
            else:
                t = t1

            # Value column: numeric
            v = pd.to_numeric(vser, errors='coerce')

            # Score this pair
            valid = t.notna() & v.notna()
            count = int(valid.sum())
            if count > best_count and count >= 5:  # require some rows
                best = (t, v)
                best_count = count

    if best is None:
        return None

    t, v = best
    df = pd.DataFrame({'time': t, 'value': v}).dropna(subset=['time', 'value'])
    if df.empty:
        return None
    return df.sort_values('time').reset_index(drop=True)


def _read_sheet_values_with_openpyxl(xlsx: Path, sheet_name: str | int | None, max_rows: int = 200, max_cols: int = 16) -> pd.DataFrame:
    """Load raw cell values via openpyxl (data_only=True) to get computed results.

    Pandas/openpyxl returns formula strings by default; using openpyxl directly with
    data_only=True allows reading cached values saved by Excel after RefreshAll.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]] if sheet_name is None else wb[sheet_name]
        rows = []
        rcount = 0
        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            rows.append(list(row))
            rcount += 1
            if rcount >= max_rows:
                break
        import pandas as pd
        df = pd.DataFrame(rows)
    finally:
        wb.close()
    return df


def _infer_unit_from_header(header: str) -> str | None:
    """Infer unit code from a header like 'PCFS K-31-01 ST_PERFORMANCE' or 'PCFS K-31 ...'."""
    m = re.search(r"\b([A-Z]+-\d{2}(?:-\d{2})?)\b", header.upper())
    return m.group(1) if m else None


def _infer_plant_unit_tag(header: str) -> tuple[str | None, str | None, str | None]:
    """Attempt to infer plant, unit, tag from a column header.

    Example: 'PCFS K-31-01 ST_PERFORMANCE' -> (PCFS, K-31-01, ST_PERFORMANCE)
    """
    s = header.strip()
    m = re.match(r"^([A-Za-z0-9_\-]+)\s+([A-Za-z]+-\d{2}(?:-\d{2})?)\s+(.+)$", s)
    if m:
        plant = m.group(1)
        unit = m.group(2)
        tag = m.group(3).strip()
        return plant, unit, tag
    # fallback: unit only via previous helper; tag as header
    unit = _infer_unit_from_header(header)
    return None, unit, header.strip()


def load_latest_frame(
    xlsx: Path,
    time_label: str = "TIME",
    value_col_hint: str = "ST_PERFORMANCE",
    unit: str | None = None,
    plant: str | None = None,
    tag: str | None = None,
    sheet_name: str | int | None = None,
) -> pd.DataFrame:
    """Load a tidy time/value DataFrame from an Excel file.

    - Finds the header row that contains the time label (default 'TIME').
    - Chooses a value column ending with value_col_hint; falls back to first non-time column.
    - Returns DataFrame with columns ['time', 'value'] sorted by time.
    """
    xlsx = Path(xlsx)

    raw = pd.read_excel(xlsx, header=None, engine="openpyxl", sheet_name=sheet_name)
    
    # Handle case where sheet_name=None returns dict of sheets
    if isinstance(raw, dict):
        # Use first sheet if multiple sheets returned
        sheet_names = list(raw.keys())
        raw = raw[sheet_names[0]]
        if sheet_name is None:
            sheet_name = sheet_names[0]  # Use the first sheet for subsequent read
    
    hdr_idx = _detect_header_row(raw)
    if hdr_idx is None:
        # Heuristic fallback for sheets that contain direct PISampDat output without headers
        parsed = _try_parse_time_value_block(raw)
        if parsed is None:
            # Try again by reading values via openpyxl with data_only=True
            try:
                raw2 = _read_sheet_values_with_openpyxl(xlsx, sheet_name)
                parsed = _try_parse_time_value_block(raw2, max_scan_cols=64)
            except Exception:
                parsed = None
        if parsed is None:
            raise RuntimeError("Header row with 'TIME' not found in first rows.")
        # Attach optional markers and return
        if unit:
            parsed["unit"] = unit
        return parsed

    df = pd.read_excel(xlsx, header=hdr_idx, engine="openpyxl", sheet_name=sheet_name)

    # map columns
    time_col = next(c for c in df.columns if str(c).strip().upper() == time_label)

    val_col = next((c for c in df.columns if str(c).strip().upper().endswith(value_col_hint)), None)
    if val_col is None:
        cand = [c for c in df.columns if c != time_col]
        if not cand:
            raise RuntimeError("No value column found.")
        val_col = cand[0]

    original_value_name = str(val_col)
    df = df.rename(columns={time_col: "time", val_col: "value"})
    df = df[[c for c in df.columns if not str(c).startswith("Unnamed")]]

    # Robust time parsing: prefer Excel serial conversion when appropriate.
    # Some sheets store timestamps as Excel serial day numbers even when the header
    # row is present. Pandas would interpret numerics as Unix seconds (1970 epoch),
    # yielding bogus dates like '1970-01-01 00:00:00.000045'. Detect and convert.
    try:
        t_raw = df["time"]
        t_str = pd.to_datetime(t_raw, errors="coerce")
        # Attempt Excel serial path
        num = pd.to_numeric(t_raw, errors="coerce")
        t_xl = pd.to_datetime(num, unit="d", origin="1899-12-30")
        c_str = t_str.notna().sum()
        c_xl = t_xl.notna().sum()
        # Heuristic: if Excel-serial yields nearly as many valid as string, prefer it
        # This matches the earlier heuristic used in _try_parse_time_value_block.
        if c_xl > 0 and c_xl >= max(1, int(c_str * 0.9)):
            df["time"] = t_xl
        else:
            df["time"] = t_str
    except Exception:
        df["time"] = pd.to_datetime(df["time"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["time", "value"]).sort_values("time").reset_index(drop=True)

    # Attach markers for plant/unit/tag so many tags can live in one dataset
    infer_plant, infer_unit, infer_tag = _infer_plant_unit_tag(original_value_name)
    if unit or plant or tag or infer_unit or infer_plant or infer_tag:
        if plant or infer_plant:
            df["plant"] = (plant or infer_plant)
        if unit or infer_unit:
            df["unit"] = (unit or infer_unit)
        if tag or infer_tag:
            df["tag"] = _slugify_tag(tag or infer_tag)
    return df


def write_parquet(df: pd.DataFrame, out_path: Path, *, engine: str | None = None) -> Path:
    """Write DataFrame to Parquet atomically with engine fallback.

    - Writes to a temp file first, then replaces the target (best‑effort atomic).
    - Prefers 'pyarrow'; falls back to 'fastparquet'.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    engines = [engine] if engine else ["pyarrow", "fastparquet"]
    last_err: Exception | None = None
    for eng in engines:
        tmp_path = out_path.with_suffix(out_path.suffix + f".tmp-{int(time.time()*1000)}")
        try:
            # Write to temp path first
            df.to_parquet(tmp_path, index=False, engine=eng)
            try:
                os.replace(tmp_path, out_path)
            except Exception as rep_err:
                # Fallback: copy over if replace is blocked, then remove tmp
                try:
                    import shutil
                    shutil.copy2(str(tmp_path), str(out_path))
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                except Exception as copy_err:
                    # Cleanup and try next engine
                    last_err = rep_err or copy_err
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                    continue
            return out_path
        except Exception as e:
            last_err = e
            # Ensure tmp cleaned up if created
            try:
                if tmp_path.exists():
                    os.remove(tmp_path)
            except Exception:
                pass
            continue

    raise RuntimeError(
        "Parquet write failed. Install 'pyarrow' (preferred) or 'fastparquet'.\n"
        f"Last error: {last_err}"
    )


def append_parquet(df: pd.DataFrame, out_path: Path) -> Path:
    """Append rows to an existing Parquet in a memory-efficient way.

    Prefers using fastparquet's append mode (does not load the target file
    into memory). Falls back to creating the file if it doesn't exist.
    If fastparquet is unavailable or append fails (e.g., schema mismatch),
    this function will raise and allow callers to choose an alternative
    strategy (such as a DuckDB union).

    Notes:
    - The DataFrame's columns are aligned to the existing file's schema.
    - Extra columns present only in the DataFrame are dropped for append.
    - Missing columns are added with NA to satisfy schema compatibility.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # If file doesn't exist, just write a new one using the robust writer
    if not out_path.exists():
        return write_parquet(df, out_path)

    try:
        import fastparquet as fp
        from fastparquet import write as fp_write

        # Ensure time dtype is consistent
        if 'time' in df.columns:
            df['time'] = pd.to_datetime(df['time'])

        # Align columns with existing schema
        pf = fp.ParquetFile(str(out_path))
        existing_cols = list(pf.columns)

        # Add any missing columns as NA
        for col in existing_cols:
            if col not in df.columns:
                df[col] = pd.NA

        # Drop extra columns that aren't in the file schema
        extra = [c for c in df.columns if c not in existing_cols]
        if extra:
            df = df.drop(columns=extra)

        # Reorder columns to match existing file
        df = df[existing_cols]

        # Append rows as a new row group
        fp_write(str(out_path), df, append=True)
        return out_path
    except Exception as e:
        # Surface the error to allow callers to switch to a different path
        raise RuntimeError(f"fastparquet append failed: {e}")
def _slugify_tag(name: str) -> str:
    t = re.sub(r"\s+", "_", name.strip())
    t = re.sub(r"[^A-Za-z0-9_\-]", "_", t)
    return t



