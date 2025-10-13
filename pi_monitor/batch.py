from __future__ import annotations

from pathlib import Path
import os
import time
from typing import Iterable

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

try:
    import xlwings as xw
except Exception:
    xw = None  # type: ignore


def _slug(s: str) -> str:
    import re
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"[^A-Za-z0-9_\-]", "_", s)
    return s


def _fetch_single(
    wb, sheet_name: str, tag: str, server: str, start: str, end: str, step: str, *, settle_seconds: float = 1.0
) -> pd.DataFrame:
    """Write a PISampDat formula for `tag` and read its spilled (time,value) array faster.

    Optimizations:
    - Prefer dynamic-array spill with `expand()` (tiny reads) and only fallback to large
      array formulas when required (older Excel).
    - Adapt target array row count to requested window/step to avoid 100k-cell reads.
    - Allow lighter calculation via EXCEL_CALC_MODE env (sheet|full), default 'sheet'.
    """
    sht = None
    try:
        sht = wb.sheets[sheet_name]
        sht.clear()
    except Exception:
        sht = wb.sheets.add(sheet_name, after=wb.sheets[-1])

    # Helper: estimate expected rows to limit large array targets when we must use them
    def _estimate_rows(start_s: str, end_s: str, step_s: str) -> int:
        import re
        from datetime import datetime, timedelta
        now = datetime.now()
        # parse end
        try:
            end_dt = now if end_s.strip() == '*' else pd.to_datetime(end_s)
        except Exception:
            end_dt = now
        # parse start (supports -Nd, -Nh, -Nm, -Ny)
        m = re.match(r"^-\s*(\d*\.?\d+)([ydhm])$", start_s.replace(" ", ""), re.IGNORECASE)
        if m:
            val = float(m.group(1))
            unit = m.group(2).lower()
            if unit == 'y':
                start_dt = end_dt - timedelta(days=365 * val)
            elif unit == 'd':
                start_dt = end_dt - timedelta(days=val)
            elif unit == 'h':
                start_dt = end_dt - timedelta(hours=val)
            else:
                start_dt = end_dt - timedelta(minutes=val)
        else:
            try:
                start_dt = pd.to_datetime(start_s)
            except Exception:
                start_dt = end_dt - timedelta(days=365)
        total_minutes = max(1.0, (end_dt - start_dt).total_seconds() / 60.0)
        # parse step (-0.1h, -5m)
        step_s = step_s.strip()
        sm = re.match(r"^-\s*(\d*\.?\d+)([hm])$", step_s, re.IGNORECASE)
        if sm:
            sval = float(sm.group(1))
            sunit = sm.group(2).lower()
            step_minutes = sval * (60.0 if sunit == 'h' else 1.0)
        else:
            # default 6 min
            step_minutes = 6.0
        rows = int((total_minutes / max(0.1, step_minutes)) + 10)  # small buffer
        # clamp
        return max(2000, min(rows, 100000))

    # Write PISampDat to A2 as a dynamic formula and let results spill.
    # Only if dynamic arrays not available, fallback to legacy array formula over a bounded target.
    tag_escaped = tag.replace('"', '""')

    # Quick connectivity probe: try PICompDat with a few server variants and select one that yields a numeric
    def _server_variants(s: str) -> list[str]:
        variants: list[str] = []
        s = (s or '').strip()
        if s:
            # alias without backslashes
            variants.append(s.lstrip('\\'))
            # alias with leading backslashes (Excel examples often show this form)
            variants.append('\\\\' + s.lstrip('\\'))
        # final fallback: workbook default server
        variants.append("")
        # dedupe while preserving order
        seen = set()
        out: list[str] = []
        for v in variants:
            if v not in seen:
                out.append(v)
                seen.add(v)
        return out

    def _probe_server_connectivity(tag_name: str, server_name: str) -> bool:
        try:
            cell = sht.range("D1")
            cell.clear()
            f = f'=PICompDat("{tag_name}","*","{server_name}")'
            cell.formula = f
            try:
                sht.book.app.api.CalculateFull()
            except Exception:
                pass
            time.sleep(1.0)
            v = cell.value
            return isinstance(v, (int, float))
        except Exception:
            return False

    # Resolve server argument. For AF attribute paths (\\AF\DB\Elem\Attr|Path),
    # let Excel use the workbook's default server by passing an empty server string.
    import os as _os_server
    is_af_path = ('\\' in tag and '|' in tag)
    base_server = (server or _os_server.getenv('PI_SERVER_NAME') or "").strip()
    chosen_server = base_server
    if not is_af_path:
        for cand in _server_variants(base_server):
            if _probe_server_connectivity(tag_escaped, cand):
                chosen_server = cand
                try:
                    print(f"[info] Using server '{cand or 'default'}' for tag {tag}")
                except Exception:
                    pass
                break
    # Ensure Excel receives double-backslash form when a server is specified
    server_literal = '' if is_af_path else chosen_server
    if server_literal and not server_literal.startswith('\\\\'):
        server_literal = '\\\\' + server_literal.lstrip('\\')
    # PISampDat behavior: When the SAME formula is in columns A and B,
    # PI DataLink returns timestamp in A and value in B (verified in working Excel files)
    # Mode=1 means timestamps ON, which gives us the 2-column behavior
    formula = f'=PISampDat("{tag_escaped}","{start}","{end}","{step}",1,"{server_literal}")'
    try:
        import os as _os_dbg_formula
        if _os_dbg_formula.getenv('DEBUG_PI_FORMULA', '').strip():
            print(f"[debug] Formula (both columns): {formula}")
    except Exception:
        pass
    max_rows = 100000  # covers ~1y at 6-min (≈87.6k)
    # Estimate rows needed
    resolved_rows = _estimate_rows(start, end, step)
    target = sht.range((2, 1), (resolved_rows + 1, 2))
    target.clear()

    # Place dynamic spill formula in A2 (preferred); fallback to array formula over target
    try:
        sht.range("A2").formula = formula
    except Exception:
        try:
            target.formula_array = formula
        except Exception:
            # Last resort: place scalar in A2 again
            sht.range("A2").formula = formula

    # Recalculate and robustly wait for PI DataLink to finish
    import os as _os
    calc_mode = _os.getenv('EXCEL_CALC_MODE', 'sheet').lower()
    # Optional overall fetch timeout (seconds). If unset, scale with settle_seconds.
    # Default increased to better tolerate slower PI servers (e.g., remote/VPN).
    try:
        _fetch_timeout = float(_os.getenv('PI_FETCH_TIMEOUT', '0').strip())
    except Exception:
        _fetch_timeout = 0.0
    if _fetch_timeout <= 0:
        _fetch_timeout = max(20.0, settle_seconds * 12.0)

    def _wait_for_datalink_completion(timeout: float) -> bool:
        t0 = time.monotonic()
        app = sht.book.app
        done_cycles = 0
        # Early error detection - fail fast on Excel errors
        early_error_detect = _os.getenv('PI_EARLY_ERROR_DETECT', '').strip() in ('1', 'true', 'yes', 'on')

        while (time.monotonic() - t0) < timeout:
            try:
                # Encourage async queries to complete if available
                try:
                    app.api.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass
                # Trigger a calculation cycle according to mode
                try:
                    if calc_mode == 'full':
                        app.api.CalculateFull()
                    else:
                        try:
                            sht.api.Calculate()
                        except Exception:
                            app.api.CalculateFull()
                except Exception:
                    pass

                # Early error detection: check for Excel error in A2 after 5s (fail fast)
                if early_error_detect and (time.monotonic() - t0) > 5.0:
                    try:
                        cell_val = sht.range("A2").value
                        if isinstance(cell_val, str) and cell_val.startswith('#'):
                            # Excel error detected - tag likely doesn't exist or server unreachable
                            print(f"[warn] Excel error detected for tag '{tag}': {cell_val} - skipping quickly")
                            return False
                    except Exception:
                        pass

                # Check Excel calculation state (XlCalculationState: 0=Done, 1=Calculating, 2=Pending)
                try:
                    state = int(app.api.CalculationState)
                except Exception:
                    state = None

                # Inspect early rows of the spill for a numeric value in second column
                vals = None
                try:
                    vals = sht.range("A2").expand().value
                except Exception:
                    vals = None
                if vals:
                    if isinstance(vals, tuple):
                        vals = list(vals)
                    if isinstance(vals, list) and vals and not isinstance(vals[0], (list, tuple)):
                        vals = [list(vals)]
                    probe = vals[: min(10, len(vals))]
                    for row in probe:
                        if isinstance(row, (list, tuple)) and len(row) >= 2:
                            v = row[1]
                            if isinstance(v, (int, float)) and pd.notna(v):
                                return True
                if state == 0:
                    done_cycles += 1
                    if done_cycles >= 3:
                        return True
                else:
                    done_cycles = 0
            except Exception:
                pass
            time.sleep(0.5)
        # Timed out
        return False

    _completed = _wait_for_datalink_completion(_fetch_timeout)
    if not _completed:
        try:
            print(f"[warn] PI DataLink completion timed out after {_fetch_timeout:.1f}s for tag '{tag}'. Forcing final calculation...")
        except Exception:
            pass
        # Force one more full calculation + wait before giving up
        try:
            app.api.CalculateFullRebuild()
            time.sleep(5)  # Give it 5 more seconds after rebuild
            app.api.CalculateFull()
            time.sleep(3)
        except Exception:
            pass

    # Read spill first (usually fastest/smallest)
    values = None
    try:
        values = sht.range("A2").expand().value
    except Exception:
        values = None

    # Debug: Log what we read
    try:
        import os as _os_debug
        if _os_debug.getenv('DEBUG_PI_FETCH', '').strip():
            rows_read = len(values) if values and isinstance(values, (list, tuple)) else 0
            print(f"[debug] Read {rows_read} rows from Excel spill for tag: {tag}")
    except Exception:
        pass

    if not values:
        # Fallback to reading the bounded target range
        values = target.value

    # If still empty, linger briefly to allow slow async completion then retry once.
    # Controlled by PI_FETCH_LINGER (seconds, default 10). Set to 0 to disable.
    def _is_empty(v) -> bool:
        if v is None:
            return True
        if isinstance(v, (list, tuple)):
            return len(v) == 0
        return False

    if _is_empty(values):
        try:
            _linger = float(_os.getenv('PI_FETCH_LINGER', '10').strip())
        except Exception:
            _linger = 10.0
        if _linger > 0:
            _wait_for_datalink_completion(_linger)
            # Retry spill then target
            try:
                values = sht.range("A2").expand().value
            except Exception:
                values = None
            if not values:
                values = target.value
    # If still empty, try to diagnose the top-left cell contents for Excel error codes
    if _is_empty(values):
        try:
            cell = sht.range("A2").value
            if isinstance(cell, str) and cell.startswith('#'):
                print(f"[warn] Excel returned error in A2 for tag {tag}: {cell}")
        except Exception:
            pass
    # Normalize to 2D list-of-lists
    if values is None:
        return pd.DataFrame(columns=["time", "value"])  # empty
    if isinstance(values, tuple):
        values = list(values)
    if isinstance(values, list) and values and not isinstance(values[0], (list, tuple)):
        # Single row spill
        values = [list(values)]
    if not isinstance(values, list):
        return pd.DataFrame(columns=["time", "value"])  # empty

    try:
        df = pd.DataFrame(values, columns=["time", "value"])  # may include blanks
    except Exception:
        df = pd.DataFrame(columns=["time", "value"])  # treat as no data

    def _normalize_df(_df: pd.DataFrame) -> pd.DataFrame:
        # PI DataLink returns Excel serial date numbers when 'Show time stamps' is on.
        # Convert Excel serial to pandas datetime using Excel's epoch.
        if _df.empty or "time" not in _df.columns or "value" not in _df.columns:
            return pd.DataFrame(columns=["time", "value"])  # empty
        try:
            ser = pd.to_numeric(_df["time"], errors="coerce")
            _df["time"] = pd.to_datetime(ser, unit="d", origin="1899-12-30")
        except Exception:
            _df["time"] = pd.to_datetime(_df["time"], errors="coerce")
        _df["value"] = pd.to_numeric(_df["value"], errors="coerce")
        _df = _df.dropna(subset=["time", "value"]).reset_index(drop=True)
        return _df

    df = _normalize_df(df)

    # If still empty, attempt a more robust second pass:
    # 1) Force full rebuild calculation
    # 2) Retry using the workbook's default PI Server (empty server string)
    if df.empty:
        try:
            app = sht.book.app
            # Full rebuild calculation to wake PI DataLink on some installs
            try:
                app.api.CalculateFullRebuild()
            except Exception:
                try:
                    app.api.CalculateFull()
                except Exception:
                    pass
            # Rewrite the formula to use the workbook's default server
            formula_default_server = f'=PISampDat("{tag_escaped}","{start}","{end}","{step}",1,"")'
            try:
                sht.range("A2").formula = formula_default_server
            except Exception:
                try:
                    target.formula_array = formula_default_server
                except Exception:
                    sht.range("A2").formula = formula_default_server

            # Wait a little longer on the second pass
            extra_wait = 0.0
            try:
                extra_wait = float(_os.getenv('PI_FETCH_SECOND_PASS', '15').strip())
            except Exception:
                extra_wait = 15.0
            if extra_wait > 0:
                t1 = time.monotonic()
                while (time.monotonic() - t1) < extra_wait:
                    try:
                        app.api.CalculateUntilAsyncQueriesDone()
                    except Exception:
                        pass
                    try:
                        sht.api.Calculate()
                    except Exception:
                        try:
                            app.api.CalculateFull()
                        except Exception:
                            pass
                    time.sleep(0.5)

            # Re-read spill
            try:
                values = sht.range("A2").expand().value
            except Exception:
                values = None
            if not values:
                values = target.value
            df = _normalize_df(pd.DataFrame(values, columns=["time", "value"]) if values else pd.DataFrame(columns=["time", "value"]))
        except Exception:
            # Keep empty df
            pass

    return df


def build_unit_from_tags(
    xlsx: Path,
    tags: Iterable[str],
    out_parquet: Path,
    *,
    plant: str,
    unit: str,
    server: str = "\\\\PTSG-1MMPDPdb01",
    start: str = "-1y",
    end: str = "*",
    step: str = "-0.1h",
    work_sheet: str = "DL_WORK",
    settle_seconds: float = 1.0,
    visible: bool = False,
    use_working_copy: bool = True,
) -> Path:
    """Loop over tags, pull 1y/0.1h via DataLink, and write a single Parquet.

    Parquet schema columns: time (timestamp[ns]), value (float64), plant, unit, tag.
    Uses a streaming ParquetWriter to avoid holding all tags in memory.
    """
    if xw is None:
        raise RuntimeError("xlwings is required for DataLink-driven fetches.")

    xlsx = Path(xlsx)
    out_parquet = Path(out_parquet)
    out_parquet.parent.mkdir(parents=True, exist_ok=True)

    # We'll create a fresh working copy per attempt

    attempts = 0
    wrote_any_rows = False
    # Try headless first; if ABF and no rows, retry with Excel visible to ensure PI DataLink loads like your manual session
    while True:
        # Create fresh working copy for this attempt (if requested)
        open_path = Path(xlsx)
        working_path = None
        if use_working_copy:
            from datetime import datetime
            import shutil
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            working_path = open_path.parent / f"{open_path.stem}_working_{ts}{open_path.suffix}"
            try:
                shutil.copy2(str(open_path), str(working_path))
                open_path = working_path
            except Exception:
                working_path = None

        app = xw.App(visible=visible, add_book=False)
        writer: pq.ParquetWriter | None = None
        rows_this_try = 0
        try:
            app.display_alerts = False
            app.screen_updating = False
            # Lower security to allow external data connections/formulas to execute without UI prompts
            try:
                app.api.AutomationSecurity = 1  # msoAutomationSecurityLow
            except Exception:
                pass
            # Best-effort: ensure PI DataLink is connected in this Excel instance
            pi_datalink_loaded = False
            try:
                for c in app.api.COMAddIns:
                    try:
                        desc = str(getattr(c, 'Description', ''))
                        prog = str(getattr(c, 'ProgId', ''))
                        name = (desc or prog).lower()
                        if ('pi' in name and 'datalink' in name) or ('pitime' in name):
                            # Force Connect multiple times - sometimes it needs coercion
                            try:
                                c.Connect = True
                                time.sleep(0.5)
                                c.Connect = True
                            except Exception:
                                pass
                            # Verify it's actually connected
                            if getattr(c, 'Connect', False):
                                pi_datalink_loaded = True
                                print(f'[info] PI DataLink COM add-in loaded: {desc or prog}')
                    except Exception:
                        pass
            except Exception:
                pass

            # Try Excel Add-ins as fallback (sometimes PI DataLink is here instead)
            if not pi_datalink_loaded:
                try:
                    for addin in app.api.AddIns:
                        try:
                            nm = str(getattr(addin, 'Name', '')).upper()
                            if 'PI' in nm and 'DATALINK' in nm:
                                addin.Installed = True
                                pi_datalink_loaded = True
                                print(f'[info] PI DataLink Excel add-in enabled: {nm}')
                        except Exception:
                            pass
                except Exception:
                    pass

            if not pi_datalink_loaded:
                print('[ERROR] ========================================')
                print('[ERROR] PI DataLink add-in NOT DETECTED!')
                print('[ERROR] All fetches will likely fail')
                print('[ERROR] Ensure PI DataLink is installed')
                print('[ERROR] ========================================')
                # Continue anyway - might work despite detection failure
            # Optional debug: list add-ins
            try:
                import os as _os_dbg
                if _os_dbg.getenv('DEBUG_XL_ADDINS', '').strip():
                    print('[excel] Listing COM Add-ins:')
                    for c in app.api.COMAddIns:
                        try:
                            print('   -', getattr(c, 'ProgId', ''), '| Connected=', getattr(c, 'Connect', None))
                        except Exception:
                            pass
                    print('[excel] Listing Excel Add-ins:')
                    for a in app.api.AddIns:
                        try:
                            print('   -', getattr(a, 'Name', ''), '| Installed=', getattr(a, 'Installed', None))
                        except Exception:
                            pass
            except Exception:
                pass
            # Also attempt to enable Excel add-ins by name, if present but not active
            try:
                for addin in app.api.AddIns:
                    try:
                        nm = str(getattr(addin, 'Name', '')).upper()
                        title = str(getattr(addin, 'Title', '')).upper()
                        if (('PI' in nm and 'DATALINK' in nm) or ('PI' in title and 'DATALINK' in title)):
                            try:
                                addin.Installed = True
                            except Exception:
                                pass
                    except Exception:
                        pass
            except Exception:
                pass

            # Open workbook with suppressed link updates and recommendations
            try:
                wb = app.books.open(str(open_path), update_links=False, read_only=False, ignore_read_only_recommended=True)
            except Exception:
                wb = app.books.open(str(open_path))

            # Warmup: Give PI DataLink time to initialize after workbook opens
            # This is critical - when Excel opens via automation, PI DataLink needs a moment to connect
            import os as _os_warmup
            warmup_seconds = float(_os_warmup.getenv('PI_DATALINK_WARMUP', '3').strip())
            if warmup_seconds > 0:
                print(f'[info] Waiting {warmup_seconds:.1f}s for PI DataLink to initialize...')
                time.sleep(warmup_seconds)
                # Trigger a calculation to wake up PI DataLink
                try:
                    app.api.CalculateFull()
                except Exception:
                    pass

            # Optional: auto-detect working PI server if requested via env
            try:
                autodetect = str(os.getenv('PI_SERVER_CANDIDATES', '')).strip()
            except Exception:
                autodetect = ''
            chosen_server = server
            if autodetect:
                try:
                    candidates = [s.strip() for s in autodetect.split(',') if s.strip()]
                    # Always try the provided server first
                    if str(server).strip():
                        if server.lstrip('\\') not in [c.lstrip('\\') for c in candidates]:
                            candidates = [server] + candidates
                    # Probe using the first non-comment tag
                    probe_tag = None
                    for t in tags:
                        ts = t.strip()
                        if ts and not ts.startswith('#'):
                            probe_tag = ts
                            break
                    if probe_tag:
                        for cand in candidates:
                            try:
                                print(f"[info] Probing PI server candidate: {cand}")
                            except Exception:
                                pass
                            df_probe = _fetch_single(
                                wb, work_sheet, probe_tag, cand, start, end, step, settle_seconds=max(1.5, settle_seconds)
                            )
                            if not df_probe.empty:
                                chosen_server = cand
                                try:
                                    print(f"[info] Selected PI server: {cand}")
                                except Exception:
                                    pass
                                break
                except Exception:
                    pass

            # Count valid tags
            valid_tags = [t.strip() for t in tags if t.strip() and not t.strip().startswith("#")]
            total_tags = len(valid_tags)
            tags_processed = 0
            tags_success = 0
            tags_nodata = 0

            for tag in tags:
                tag = tag.strip()
                if not tag or tag.startswith("#"):
                    continue

                tags_processed += 1
                print(f"[{tags_processed}/{total_tags}] Fetching: {tag}")

                df = _fetch_single(wb, work_sheet, tag, chosen_server, start, end, step, settle_seconds=settle_seconds)
                if df.empty:
                    tags_nodata += 1
                    print(f"[warn] No data for tag: {tag}")
                    print(f"[progress] Success: {tags_success}/{total_tags} | No data: {tags_nodata}/{total_tags} | Remaining: {total_tags - tags_processed}")
                    continue

                tags_success += 1
                df["plant"] = plant
                df["unit"] = unit
                df["tag"] = _slug(tag)
                print(f"[progress] Success: {tags_success}/{total_tags} | No data: {tags_nodata}/{total_tags} | Remaining: {total_tags - tags_processed}")

                # Reorder columns
                df = df[["time", "value", "plant", "unit", "tag"]]

                table = pa.Table.from_pandas(df, preserve_index=False)
                if writer is None:
                    writer = pq.ParquetWriter(str(out_parquet), table.schema, compression="zstd")
                writer.write_table(table)
                rows_this_try += len(df)

            # Print final summary
            print(f"\n{'='*70}")
            print(f"FETCH SUMMARY FOR {unit}")
            print(f"{'='*70}")
            print(f"Total tags:       {total_tags}")
            print(f"Success:          {tags_success} ({tags_success*100//total_tags if total_tags > 0 else 0}%)")
            print(f"No data:          {tags_nodata} ({tags_nodata*100//total_tags if total_tags > 0 else 0}%)")
            print(f"Total rows fetched: {rows_this_try:,}")
            print(f"{'='*70}\n")

            try:
                wb.save()
            except Exception:
                pass
            wb.close()
        finally:
            if writer is not None:
                writer.close()
            app.quit()
            # Clean up the working copy if used
            if working_path is not None:
                try:
                    working_path.unlink(missing_ok=True)
                except Exception:
                    pass

        wrote_any_rows = wrote_any_rows or (rows_this_try > 0)
        attempts += 1
        # Retry rule: if initial headless attempt wrote 0 rows for any plant,
        # retry once with Excel visible to ensure PI DataLink loads correctly.
        # Allow skipping the visible Excel fallback via env NO_VISIBLE_FALLBACK=1|true|yes|on
        if ((not visible) and rows_this_try == 0 and attempts == 1):
            no_visible = str(os.getenv('NO_VISIBLE_FALLBACK', '')).strip().lower() in {"1", "true", "yes", "on"}
            if no_visible:
                try:
                    print("[info] Headless fetch returned no rows; skipping visible Excel fallback due to NO_VISIBLE_FALLBACK.")
                except Exception:
                    pass
                break
            try:
                print("[info] Headless fetch returned no rows; retrying with Excel visible to ensure PI DataLink loads...")
            except Exception:
                pass
            visible = True
            continue
        break

    return out_parquet


def read_tags_from_sheet(xlsx: Path, sheet_name: str = "DL_K12_01", row: int = 2, max_cols: int = 256) -> list[str]:
    """Read tags from a single row in an Excel sheet (A<Row>.. up to blank).

    This is useful when you already maintain a tag list in the workbook.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name]
    out: list[str] = []
    for c in range(1, max_cols + 1):
        v = ws.cell(row=row, column=c).value
        if v is None or str(v).strip() == "":
            if c > 1 and (ws.cell(row=row, column=c + 1).value is None):
                break
            continue
        s = str(v).strip()
        # Heuristic: PI tags often contain dots, no spaces, not starting with '-' and not a UNC path
        if "." in s and " " not in s and not s.startswith("-") and not s.startswith("\\\\") and any(ch.isalpha() for ch in s):
            out.append(s)
    wb.close()
    return out
