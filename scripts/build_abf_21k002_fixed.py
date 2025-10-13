#!/usr/bin/env python3
"""
ABF 21-K002 (C2 Compressor) Parquet builder.
Uses AF (Asset Framework) path format: \\server\plant\unit\equipment\parameter|attribute
"""
from __future__ import annotations

from pathlib import Path
import sys
import os

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from pi_monitor.batch import build_unit_from_tags  # noqa: E402
from pi_monitor.clean import dedup_parquet  # noqa: E402


def read_tag_paths(xlsx_path: Path) -> list[str]:
    """Extract AF paths from Excel: Column A (equipment group) + Column B (parameter name)"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    base_path = r"\\VCENPOCOEPTNAP01\Protean\ABF\Compressor\21-K002"
    tags = []

    for row in range(2, ws.max_row + 1):
        equipment_group = ws.cell(row, 1).value  # Column A: "C2 - Comp-Comp"
        parameter_name = ws.cell(row, 2).value    # Column B: "TURB_TBRG_T_A"

        if equipment_group and parameter_name:
            # Build full AF path
            af_path = f"{base_path}\\{equipment_group}\\{parameter_name}|OperatingValue"
            tags.append(af_path)

    return tags


def main() -> int:
    xlsx = PROJECT_ROOT / "excel" / "ABFSB" / "ABF LIMIT REVIEW (CURRENT).xlsx"
    safe_unit = "21-K002"
    out_parquet = PROJECT_ROOT / "data" / "processed" / f"{safe_unit}_1y_0p1h.parquet"

    plant = "ABFSB"
    unit = "21-K002"

    # For AF paths, pass empty server - Excel will use workbook's default
    server = ""  # AF paths contain full server path already
    start = "-1y"
    end = "*"
    step = "-0.1h"
    work_sheet = "DL_WORK"
    settle_seconds = 2.0
    visible = True

    print(f"Extracting AF paths from Excel...")
    tags = read_tag_paths(xlsx)

    if not tags:
        raise SystemExit(f"No tags found in {xlsx}")

    print(f"Building Parquet for {plant} {unit} with {len(tags)} AF tags...")
    print(f"Sample tag: {tags[0]}")
    print(f"Start time: {start}, End: {end}, Step: {step}")

    out = build_unit_from_tags(
        xlsx,
        tags,
        out_parquet,
        plant=plant,
        unit=unit,
        server=server,  # Empty for AF paths
        start=start,
        end=end,
        step=step,
        work_sheet=work_sheet,
        settle_seconds=settle_seconds,
        visible=visible,
    )

    print(f"Wrote: {out}")

    dedup = dedup_parquet(out)
    print(f"Master (dedup) ready: {dedup}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
