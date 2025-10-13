#!/usr/bin/env python3
"""
Sample script to fetch PI data for ABF plant using Sheet1 in ABFSB Excel file.
This demonstrates how to read existing data and fetch new data using PI DataLink.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys
from datetime import datetime

def read_existing_abf_data(excel_path: str) -> pd.DataFrame:
    """Read existing ABF data from Sheet1 in the Excel file."""
    print(f"Reading existing ABF data from: {excel_path}")
    
    # Load the workbook and Sheet1
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']
    
    # Read data into pandas DataFrame
    data = []
    for row in range(2, min(100, ws.max_row + 1)):  # Read first 100 rows as sample
        timestamp = ws.cell(row, 1).value
        value = ws.cell(row, 2).value
        if timestamp and value is not None:
            data.append({'time': timestamp, 'value': value})
    
    wb.close()
    
    df = pd.DataFrame(data)
    if not df.empty:
        df['time'] = pd.to_datetime(df['time'])
        df = df.sort_values('time').reset_index(drop=True)
    
    return df

def fetch_new_abf_data_sample():
    """Sample function to demonstrate fetching new ABF data using PI DataLink."""
    # Configuration for ABF data fetching
    config = {
        'excel_file': 'excel/ABFSB/ABFSB_Automation_Master.xlsx',
        'plant': 'ABF',
        'unit': '07-MT01-K001',
        'sample_tag': 'ABF.07-MT001.FI-07054.PV',  # Tag from Sheet1
        'server': r"\\PTSG-1MMPDPdb01",  # Default server
        'start': "-24h",  # Last 24 hours
        'end': "*",       # Current time
        'step': "-0.1h",  # 6-minute intervals
        'work_sheet': "Sheet1"  # Use Sheet1 for fetching
    }
    
    print("ABF Data Fetching Configuration:")
    for key, value in config.items():
        print(f"  {key}: {value}")
    
    return config

def analyze_existing_data(df: pd.DataFrame):
    """Analyze the existing ABF data structure."""
    if df.empty:
        print("No data found in Sheet1")
        return
    
    print(f"\nData Analysis:")
    print(f"Total rows: {len(df)}")
    print(f"Time range: {df['time'].min()} to {df['time'].max()}")
    print(f"Value statistics:")
    print(f"  Min: {df['value'].min():.2f}")
    print(f"  Max: {df['value'].max():.2f}")
    print(f"  Mean: {df['value'].mean():.2f}")
    print(f"  Latest value: {df['value'].iloc[-1]:.2f} at {df['time'].iloc[-1]}")

def main():
    """Main function to demonstrate ABF data fetching."""
    excel_path = "excel/ABFSB/ABFSB_Automation_Master.xlsx"
    
    # Check if file exists
    if not Path(excel_path).exists():
        print(f"Error: Excel file not found at {excel_path}")
        return 1
    
    try:
        # 1. Read existing data from Sheet1
        existing_data = read_existing_abf_data(excel_path)
        
        # 2. Analyze the data
        analyze_existing_data(existing_data)
        
        # 3. Show configuration for new data fetching
        config = fetch_new_abf_data_sample()
        
        # 4. Demonstrate how to use the build script
        print(f"\nTo fetch new ABF data, run:")
        print(f"python scripts/build_abf_07mt01_k001.py")
        print(f"\nOr manually fetch using:")
        print(f"from pi_monitor.batch import build_unit_from_tags")
        print(f"from pathlib import Path")
        print(f"tags = ['ABF.07-MT001.FI-07054.PV']  # Sample tag")
        print(f"build_unit_from_tags(...)")
        
        return 0
        
    except Exception as e:
        print(f"Error: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main())